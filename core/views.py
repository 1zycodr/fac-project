from django.http import HttpResponse, HttpResponseRedirect
from django.views import View
from django.views.generic import (
	ListView, CreateView, DeleteView, UpdateView, DetailView
)
from core import models, forms
import xlwt
import os
import openpyxl
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.conf import settings


class FACListView(ListView):
	model = models.FuncAcceptanceCertificate
	queryset = models.FuncAcceptanceCertificate.objects.all().order_by('date')
	template_name = 'fac/index.html'


class SystemListView(ListView):
	model = models.SystemList
	queryset = models.SystemList.objects.all().order_by('date')
	template_name = 'system/index.html'


class FACDeleteView(DeleteView):
	model = models.FuncAcceptanceCertificate
	success_url = '/certificates'

	def get(self, request, *args, **kwargs):
		return self.post(request, *args, **kwargs)


class FACCreateView(CreateView):
	model = models.FuncAcceptanceCertificate
	form_class = forms.CreateFuncAcceptanceCertificate
	template_name = 'fac/create.html'
	success_url = '/certificates'


class FACEditView(UpdateView):
	model = models.FuncAcceptanceCertificate
	form_class = forms.EditFuncAcceptanceCertificate
	template_name = 'fac/edit.html'
	success_url = '/certificates'


class FACExportView(ListView):
	model = models.FuncAcceptanceCertificate
	queryset = models.FuncAcceptanceCertificate.objects.all().order_by('date')

	def get(self, request, *args, **kwargs):
		response = HttpResponse(content_type='application/ms-excel')
		response['Content-Disposition'] = 'attachment; filename="file.xls"'
		wb = xlwt.Workbook(encoding='utf-8')
		ws = wb.add_sheet('Функционально - приёмочные акты')
		cols_widths = {}
		row_num = 0

		font_style = xlwt.XFStyle()
		font_style.font.bold = True

		cols = [
			'Номер', 'Протокол тестирования', 'Описание патча',
			'Объекты', 'Дата посадки', 'Заказчик', 'Исполнитель',
		]
		for col_num, col_title in enumerate(cols):
			cols_widths[col_num] = len(col_title) + 1
			ws.write(row_num, col_num, col_title, font_style)

		font_style = xlwt.XFStyle()
		rows = models.FuncAcceptanceCertificate.objects.all()
		for row in rows:
			row_num += 1
			ws.write(row_num, 0, row.number, font_style)
			cols_widths[0] = max(cols_widths.get(0, 1), len(row.number))
			ws.write(row_num, 1, row.protocol, font_style)
			cols_widths[1] = max(cols_widths.get(1, 1), len(row.protocol))
			ws.write(row_num, 2, row.description, font_style)
			cols_widths[2] = max(cols_widths.get(2, 1), len(row.description))
			ws.write(row_num, 3, row.objs, font_style)
			cols_widths[3] = max(cols_widths.get(3, 1), len(row.objs))
			ws.write(row_num, 4, row.date.strftime('%d.%m.%Y'), font_style)
			cols_widths[4] = max(cols_widths.get(4, 1), len(row.date.strftime('%d.%m.%Y')))
			ws.write(row_num, 5, row.customer, font_style)
			cols_widths[5] = max(cols_widths.get(5, 1), len(row.customer))
			ws.write(row_num, 6, row.executor, font_style)
			cols_widths[6] = max(cols_widths.get(6, 1), len(row.executor))
		for col_num, value in cols_widths.items():
			ws.col(col_num).width = (value + 1) * 256
		wb.save(response)
		return response


class FACUploadView(View):

	def post(self, request, *args, **kwargs):
		try:
			file = request.FILES.get('file')
			path = default_storage.save('tmp/' + str(file), ContentFile(file.read()))
			tmp_file = os.path.join(settings.MEDIA_ROOT, path)
			workbook = openpyxl.load_workbook(tmp_file)
			ws = workbook.active
			list_create = []
			for i in range(1, ws.max_row):
				name, type, date = ws.iter_cols(1, 3)
				print(str(date[i].value))
				list_create.append(
					models.SystemList(name=str(name[i].value),
									  type=str(type[i].value),
									  date=str(date[i].value.strftime('%Y-%m-%d')))
				)
			models.SystemList.objects.all().delete()
			models.SystemList.objects.bulk_create(list_create)
		except Exception as ex:
			print(ex)
		return HttpResponseRedirect('/system-list')


class FACExportDetailsView(DetailView):
	model = models.FuncAcceptanceCertificate
	queryset = models.FuncAcceptanceCertificate.objects.all().order_by('date')

	def get(self, request, *args, **kwargs):
		response = HttpResponse(content_type='application/ms-excel')
		response['Content-Disposition'] = 'attachment; filename="file.xls"'
		wb = xlwt.Workbook(encoding='utf-8')
		ws = wb.add_sheet('Функционально - приёмочный акт')
		font_style = xlwt.XFStyle()
		obj = self.get_object(self.get_queryset())

		ws.write(0, 1, 'ТОО "DBK"', font_style)
		ws.write(2, 0, 'г. Алматы', font_style)
		ws.write(2, 6, f'Дата: {obj.date}', font_style)
		ws.write(4, 1, 'Функционально-приемочный акт АБС "Colvir"', font_style)
		ws.write(5, 1, f'ФПА №{obj.number}, {obj.protocol}', font_style)
		ws.write_merge(
			7, 8, 0, 13,
			'Настоящий протокол составлен по результатам тестирования '
			'и внедрения обновлений на базе Автоматической Банковской Системы "Colvir".',
			font_style
		)
		ws.write(9, 0, 'Заказчик:', font_style)
		ws.write(9, 1, obj.customer, font_style)
		ws.write(10, 0, 'Описание обновления:', font_style)
		ws.write(10, 1, obj.description, font_style)
		ws.write(11, 0, 'Тестирование выполнил:', font_style)
		ws.write(11, 1, obj.executor, font_style)
		ws.write(12, 0, 'Объекты:', font_style)
		ws.write(12, 1, obj.objs, font_style)
		ws.write(13, 0, 'Дата посадки обновлений на продуктивную среду:', font_style)
		ws.write(13, 1, str(obj.date), font_style)
		ws.col(0).width = (len('Дата посадки обновлений на продуктивную среду:') + 1) * 256
		ws.col(1).width = (len('Дата посадки обновлений') + 1) * 256
		wb.save(response)
		return response


class FACCheckView(ListView):
	model = models.SystemList
	template_name = 'system/check.html'
	queryset = models.SystemList.objects.order_by('date')

	def get_queryset(self):
		qs = super(FACCheckView, self).get_queryset()
		objects = models.FuncAcceptanceCertificate.objects.all()
		objects_names = qs.values_list('name', flat=True)
		add_names = set()
		exclude_names = set()
		for obj in objects:
			for name in obj.objs.replace(' ', '').split(','):
				if name in objects_names:
					for f in qs:
						if f.name == name:
							if obj.date == f.date:
								exclude_names.add(name)

		return qs.exclude(name__in=exclude_names)
